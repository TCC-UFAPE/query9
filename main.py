import os
import json
from groq import Groq
import sys
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    with open("config.json", "r", encoding="utf-8") as config_file:
        config = json.load(config_file)
        api_key = config["api_key"]
except FileNotFoundError:
    print("Erro: Arquivo 'config.json' não encontrado. Crie este arquivo com sua chave da API Groq.")
    print("Exemplo de config.json: {\"api_key\": \"SUA_CHAVE_API_AQUI\"}")
    sys.exit(1)
except KeyError:
    print("Erro: Chave 'api_key' não encontrada no arquivo 'config.json'.")
    sys.exit(1)

client = Groq(api_key=api_key)

CODE_FILE_EXTENSIONS = {
    '.cpp', '.c', '.h', '.hpp', '.java', '.py', '.js', '.ts', '.cc',
    '.html', '.css', '.go', '.rs', '.php', '.rb', '.swift', '.kt'
}

def remove_comments(code):
    code = re.sub(r'/\*.*?\*/', '', code, flags=re.DOTALL)
    code = re.sub(r'(//|#).*', '', code)
    code = "\n".join(line.strip() for line in code.splitlines() if line.strip())
    return code

def analyze_code_files(root_dir, model_name, output_excel_path):
    all_results_for_report = []

    print(f"Iniciando análise dos arquivos em: {root_dir}")
    print(f"Utilizando o modelo de IA: {model_name}")
    print(f"Procurando por arquivos com as extensões: {', '.join(CODE_FILE_EXTENSIONS)}")

    for subdir, _, files in os.walk(root_dir):
        for filename in files:
            if any(filename.endswith(ext) for ext in CODE_FILE_EXTENSIONS):
                filepath = os.path.join(subdir, filename)
                relative_filepath = os.path.relpath(filepath, root_dir)
                print(f"\nProcessando: {relative_filepath}")

                try:
                    with open(filepath, "r", encoding="utf-8", errors='ignore') as f:
                        original_code = f.read()

                    if not original_code.strip():
                        print(f"   -> Aviso: Arquivo '{relative_filepath}' está vazio.")
                        all_results_for_report.append({
                            "filename": relative_filepath,
                            "code": "O arquivo original estava vazio.",
                            "analysis": "Não analisado (arquivo vazio)."
                        })
                        continue

                    cleaned_code = remove_comments(original_code)

                    if not cleaned_code.strip():
                        print(f"   -> Aviso: Arquivo '{relative_filepath}' ficou vazio após a remoção de comentários.")
                        all_results_for_report.append({
                            "filename": relative_filepath,
                            "code": original_code,
                            "analysis": "Não analisado (vazio após remoção de comentários)."
                        })
                        continue

                    system_message_content = (
                        "You are a security researcher specialized in detecting security vulnerabilities.\n"
                        "Provide the answer only in the following format:\n\n"
                        "vulnerability: <YES or NO> | vulnerability type: <type or N/A> | vulnerability name: <name or N/A> | explanation: <explanation for the prediction>.\n"
                        "Do not include anything else in the response."
                    )
                    user_message_content = (
                        "User: Is this code snippet subject to any security vulnerability?\n\n"
                        f"{cleaned_code}\n\n"
                        "Answer:"
                    )

                    print(f"   -> Enviando '{relative_filepath}' para análise...")
                    chat_completion = client.chat.completions.create(
                        messages=[
                            {"role": "system", "content": system_message_content},
                            {"role": "user", "content": user_message_content},
                        ],
                        model=model_name,
                        temperature=0,
                    )
                    analysis_result = chat_completion.choices[0].message.content
                    print(f"   -> Análise recebida para '{relative_filepath}'.")

                    all_results_for_report.append({
                        "filename": relative_filepath,
                        "code": cleaned_code,
                        "analysis": analysis_result.strip()
                    })

                except Exception as e:
                    print(f"   -> ERRO ao processar {relative_filepath}: {e}")
                    all_results_for_report.append({
                        "filename": relative_filepath,
                        "code": "Erro durante a leitura ou processamento do arquivo.",
                        "analysis": f"Erro: {e}"
                    })

    print(f"\nGerando relatório Excel em: {output_excel_path}")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Vulnerabilidades"

    title_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=12, bold=True)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    sheet.merge_cells('A1:C1')
    title_cell = sheet['A1']
    title_cell.value = f"Relatório de Análise de Vulnerabilidades (Modelo IA: {model_name})"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = title_fill
    sheet.row_dimensions[1].height = 30

    headers = ["Arquivo", "Código Limpo Analisado", "Resultado da Análise"]
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num, value=header_title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    sheet.row_dimensions[2].height = 25

    current_row = 3
    if not all_results_for_report:
        sheet.cell(row=current_row, column=1, value="Nenhum arquivo de código foi encontrado ou processado.")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    else:
        for result in all_results_for_report:
            data_to_write = [
                result['filename'],
                result['code'].strip(),
                result['analysis']
            ]
            for col_num, cell_value in enumerate(data_to_write, 1):
                cell = sheet.cell(row=current_row, column=col_num, value=cell_value)
                cell.alignment = cell_alignment
                cell.border = thin_border
            current_row += 1

    sheet.column_dimensions[get_column_letter(1)].width = 40
    sheet.column_dimensions[get_column_letter(2)].width = 70
    sheet.column_dimensions[get_column_letter(3)].width = 70

    try:
        workbook.save(output_excel_path)
        print(f"Relatório Excel gerado com sucesso: {output_excel_path}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")

if __name__ == "__main__":
    repo_root_directory = os.path.join("systemd11", "src", "vconsole")

    if not os.path.isdir(repo_root_directory):
        print(f"Erro: Diretório base para os códigos não encontrado: '{repo_root_directory}'")
        print(f"Verifique se o caminho está correto em relação à localização do script '{os.path.basename(__file__)}'.")
    else:
        GROQ_MODELS_TO_USE = [
            "moonshotai/kimi-k2-instruct-0905",
            "qwen/qwen3-32b",
            "gemma2-9b-it",
            "meta-llama/Llama-Guard-4-12B",
            "llama-3.3-70b-versatile",
            "llama-3.1-8b-instant",
            "meta-llama/llama-4-maverick-17b-128e-instruct",
            "meta-llama/llama-4-scout-17b-16e-instruct",
        ]

        for model in GROQ_MODELS_TO_USE:
            print(f"--- Executando análise para o modelo: {model} ---")
            
            folder_name = os.path.basename(repo_root_directory)
            
            sanitized_model_name = model.replace("/", "-")
            
            output_filename = f"{folder_name} - {sanitized_model_name}.xlsx"
            
            output_report_path = os.path.join(repo_root_directory, output_filename)

            analyze_code_files(repo_root_directory, model, output_report_path)
            
            print(f"--- Análise para o modelo: {model} concluída ---\n")


# GROQ_MODELS_TO_USE = [
#     "moonshotai/kimi-k2-instruct-0905", -- OK
#     "qwen/qwen3-32b", -- OK
#     "gemma2-9b-it", -- OK
#     "meta-llama/Llama-Guard-4-12B", -- OK
#     "llama-3.3-70b-versatile", --OK
#     "llama-3.1-8b-instant", -- OK
#     "meta-llama/llama-4-maverick-17b-128e-instruct",  -- OK
#     "meta-llama/llama-4-scout-17b-16e-instruct",  -- OK
#     "llama3-70b-8192", -- OK -- B.O
#     "llama3-8b-8192", -- OK -- B.O
#     "deepseek-r1-distill-llama-70b", -- OK -- B.O
#     "qwen-qwq-32b", -- OK -- B.O
#     "llama-guard-3-8b", -- OK -- B.O
#     "playai-tts", -- B.O
#     "metal-lhama/lhama-prompt-guard-2-22m", -- B.O
#     "metal-lhama/lhama-prompt-guard-2-86m", -- B.O
#     "compound-beta", -- REMOVIDAS
#     "compound-beta-mini" -- REMOVIDAS
#     "mistral-saba-24b", -- APENAS PARA OUTRAS LINGUAS
# ]